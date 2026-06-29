#!/usr/bin/env python3
# Build enforcement.json from the SFMTA records-request citations (GPS-geocoded, request #26-5453).
# Matches each ticket to the nearest CNN street segment (<=40m) instead of the lossy address join,
# then aggregates by (cnn, jsDow) -> [n, avgMin, loMin, hiMin], same schema the app already reads.
import openpyxl, json, math, os, sys, itertools, datetime
import urllib.request, urllib.parse
import numpy as np
from shapely import STRtree, points, LineString

XLSX = os.environ.get("CURB_CITATIONS_XLSX", "/Users/alevizio/Downloads/TRC7.2.22_1.1.2024_06.24.2026.xlsx")  # SFMTA records #26-5453 (not redistributable); set CURB_CITATIONS_XLSX to override
OUT  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "enforcement.json")
TOL_M = 40.0          # max citation->segment distance (meters)
MIN_N = 5
SFLAT, SFLON = (37.69, 37.84), (-122.53, -122.34)
LAT0, LON0 = 37.7749, -122.4194
KX = math.cos(math.radians(LAT0)) * 111320.0
KY = 110540.0
def proj(lon, lat): return ((lon - LON0) * KX, (lat - LAT0) * KY)
DAY = {'sun':0,'mon':1,'tue':2,'wed':3,'thu':4,'fri':5,'sat':6}

def fetch(url, params):
    u = url + '?' + urllib.parse.urlencode(params)
    req = urllib.request.Request(u, headers={'User-Agent': 'curb-enf-records'})
    with urllib.request.urlopen(req, timeout=90) as r:
        return json.load(r)

# 1) sweep segments + schedule (DataSF yhqp-riqs) -------------------------------------------------
print("fetching sweep segments…", file=sys.stderr)
rows = fetch('https://data.sfgov.org/resource/yhqp-riqs.json',
             {'$select': 'cnn,weekday,fromhour,tohour,line', '$limit': '45000'})
seg_line, sched = {}, {}
for r in rows:
    cnn = r.get('cnn')
    if not cnn: continue
    wd = r.get('weekday')
    if wd:
        dow = DAY.get(wd.strip().lower()[:3])
        try: fromH = int(float(r.get('fromhour')))
        except (TypeError, ValueError): fromH = None
        if dow is not None and fromH is not None:
            try: toH = int(float(r.get('tohour')))
            except (TypeError, ValueError): toH = fromH + 2
            sched.setdefault(cnn, {})[dow] = (fromH * 60, toH * 60)
    if cnn not in seg_line:
        ln = r.get('line')
        if ln and ln.get('coordinates'):
            try: seg_line[cnn] = LineString([proj(c[0], c[1]) for c in ln['coordinates']])
            except Exception: pass
cnns = list(seg_line.keys())
tree = STRtree([seg_line[c] for c in cnns])
print(f"  {len(cnns)} segments w/ geometry, {len(sched)} w/ schedule", file=sys.stderr)

# 2) stream citations -> arrays ------------------------------------------------------------------
print("reading citations…", file=sys.stderr)
wb = openpyxl.load_workbook(XLSX, read_only=True)
xs, ys, mins, dows = [], [], [], []
for sh in wb.sheetnames:
    it = wb[sh].iter_rows(values_only=True)
    first = next(it, None)
    if first is None: continue
    seq = it if (first and str(first[0]).strip().lower() == 'ticket number') else itertools.chain([first], it)
    for r in seq:
        if not r or r[0] is None: continue
        lat, lon, d, t = r[6], r[7], r[1], r[2]
        if lat is None or lon is None or d is None or t is None: continue
        try: lat, lon = float(lat), float(lon)
        except (TypeError, ValueError): continue
        if not (SFLAT[0] <= lat <= SFLAT[1] and SFLON[0] <= lon <= SFLON[1]): continue
        if not hasattr(t, 'hour'): continue
        x, y = proj(lon, lat)
        xs.append(x); ys.append(y)
        mins.append(t.hour * 60 + t.minute)
        dows.append((d.weekday() + 1) % 7)   # python Mon0..Sun6 -> JS getDay Sun0..Sat6
xs = np.asarray(xs); ys = np.asarray(ys)
mins = np.asarray(mins, dtype=np.int32); dows = np.asarray(dows, dtype=np.int8)
print(f"  {len(xs)} citations with coords+time in SF", file=sys.stderr)

# 3) nearest-segment match (vectorized, chunked) + aggregate -------------------------------------
acc = {}                                   # (cnn,dow) -> [n,sum,lo,hi]
bdoff = {}                                 # (cnn,dow) -> list of minutes-into-window offsets
hist = np.zeros(400, dtype=np.int64)       # minutes-into-window offset, index = off+100
matched = 0
CH = 200000
for s in range(0, len(xs), CH):
    e = min(s + CH, len(xs))
    res, dist = tree.query_nearest(points(xs[s:e], ys[s:e]), return_distance=True, all_matches=False)
    in_idx, tr_idx = res[0], res[1]
    for j in range(len(in_idx)):
        if dist[j] > TOL_M: continue
        gi = int(in_idx[j]) + s
        cnn = cnns[int(tr_idx[j])]
        sc = sched.get(cnn)
        if not sc: continue
        dow = int(dows[gi]); minute = int(mins[gi])
        if dow not in sc: continue
        fromMin, toMin = sc[dow]
        if minute < fromMin - 60 or minute > toMin + 120: continue
        matched += 1
        hist[min(max(minute - fromMin + 100, 0), 399)] += 1
        bdoff.setdefault((cnn, dow), []).append(minute - fromMin)
        a = acc.get((cnn, dow))
        if a is None: a = [0, 0, 1440, 0]; acc[(cnn, dow)] = a
        a[0] += 1; a[1] += minute
        a[2] = min(a[2], max(minute, fromMin - 30))
        a[3] = max(a[3], min(minute, toMin + 60))
    print(f"  matched {matched} (through {e})", file=sys.stderr)

# 4) output + stats ------------------------------------------------------------------------------
out = {}; kept = 0; blocks = set(); blockday_off = []
for (cnn, dow), a in acc.items():
    if a[0] < MIN_N: continue
    avg = round(a[1] / a[0])
    out.setdefault(cnn, {})[str(dow)] = [a[0], avg, a[2], a[3]]
    kept += 1; blocks.add(cnn)
    blockday_off.append(avg - sched[cnn][dow][0])

tot = int(hist.sum()); cdf = np.cumsum(hist)
within = lambda m: 100.0 * cdf[min(max(m + 100, 0), 399)] / tot
median_off = int(np.searchsorted(cdf, tot / 2) - 100)
bd = np.array(blockday_off)
print("\n=== RESULTS ===")
print(f"matched tickets: {tot:,}  ({100*tot/len(xs):.1f}% of SF citations)")
print(f"blocks: {len(blocks):,} | side-days: {kept:,}")
print(f"median ticket: {median_off} min into the window")
print(f"within window-start +22min: {within(22):.1f}%  | +45min: {within(45):.1f}%  | +60min: {within(60):.1f}%")
print(f"median block-day avg: {int(np.median(bd))} min into window  (mean {bd.mean():.1f})")

# launch-kit-style per-block-side metrics (the "22-minute window" + "87% within 45min" claims)
p50s, spans, p90le45, maxle45, span_le22 = [], [], 0, 0, 0
nb = 0
for (cnn, dow), offs in bdoff.items():
    if len(offs) < MIN_N: continue
    nb += 1
    a = np.array(offs)
    p10, p50, p90 = np.percentile(a, [10, 50, 90])
    p50s.append(p50); spans.append(p90 - p10)
    if p90 <= 45: p90le45 += 1
    if a.max() <= 45: maxle45 += 1
    if (p90 - p10) <= 22: span_le22 += 1
print("\n=== launch-kit-style metrics (per block-side) ===")
print(f"typical block-side: median ticket {int(np.median(p50s))} min into window; "
      f"middle-80% ticket span {int(np.median(spans))} min (p10-p90)")
print(f"block-sides where 90% of tickets fall within 45 min of start: {100*p90le45/nb:.1f}%")
print(f"block-sides where ALL tickets fall within 45 min of start:    {100*maxle45/nb:.1f}%")
print(f"block-sides with a <=22-min middle-80% span:                  {100*span_le22/nb:.1f}%")

out["_meta"] = {
    "generated": datetime.datetime.now().isoformat(timespec="seconds"),
    "window_since": "2024-01-02", "window_until": "2026-06-24", "min_samples": MIN_N,
    "blocks": len(blocks), "side_days": kept, "matched_citations": tot,
    "source": "SFMTA public-records request #26-5453 (TRC7.2.22 street-cleaning citations, GPS-geocoded) joined to yhqp-riqs (segments + schedule)",
    "method": "each citation GPS-matched to nearest CNN segment (<=40m), aggregated by cnn x jsDow",
    "note": "avgMin/loMin/hiMin are local minutes-of-day; dow is JS getDay (0=Sun).",
}
json.dump(out, open(OUT, "w"))
import os
print(f"wrote {OUT}  ({os.path.getsize(OUT)//1024} KB)")
